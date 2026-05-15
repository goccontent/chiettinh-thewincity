# -*- coding: utf-8 -*-
"""
App tra giá theo Mã sản phẩm từ file "GIỎ HÀNG" và tạo bản CHIETTINH cho từng căn.

Cách dùng:
    python app_chiettinh.py
"""
import os
import re
import shutil
import sys
import subprocess
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
GIO_HANG_FILE = BASE_DIR / "GIỎ HÀNG LM VAT THE WIN CITY 25.04.2026 (1).xlsx"
CHIETTINH_FILE = BASE_DIR / "CHIETTINH.xlsx"
OUTPUT_DIR = BASE_DIR / "ChietTinh_DaTao"


def parse_price(v):
    """Giỏ hàng có ô lưu số, có ô lưu chuỗi '2.098.000.000'. Chuyển hết về int."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    digits = re.sub(r"[^\d]", "", s)
    return int(digits) if digits else None


def load_products():
    """Đọc file Giỏ hàng → dict[mã căn] = thông tin."""
    if not GIO_HANG_FILE.exists():
        raise FileNotFoundError(f"Không tìm thấy file: {GIO_HANG_FILE}")
    wb = openpyxl.load_workbook(GIO_HANG_FILE, data_only=True)
    ws = wb["Sheet1"]
    products = {}
    for r in range(3, ws.max_row + 1):
        ma = ws.cell(r, 2).value
        if not ma:
            continue
        ma = str(ma).strip()
        products[ma] = {
            "ma": ma,
            "loai": ws.cell(r, 3).value,
            "so_can": ws.cell(r, 4).value,
            "so_tang": ws.cell(r, 5).value,
            "thap": ws.cell(r, 6).value,
            "dt_thong_thuy": ws.cell(r, 7).value,
            "dt_tim_tuong": ws.cell(r, 8).value,
            "huong": ws.cell(r, 10).value,
            "view": ws.cell(r, 11).value,
            "gia_vn": parse_price(ws.cell(r, 12).value),
            "gia_nn": parse_price(ws.cell(r, 13).value),
            "tinh_trang": ws.cell(r, 14).value,
        }
    return products


def fmt_vnd(n):
    if n is None:
        return ""
    return f"{n:,.0f}".replace(",", ".") + " ₫"


def create_chiettinh(product, is_foreigner, block_value):
    """Sao chép file CHIETTINH, điền PL1A!C7 = mã, C11 = giá, rồi recalc."""
    if not CHIETTINH_FILE.exists():
        raise FileNotFoundError(f"Không tìm thấy file: {CHIETTINH_FILE}")
    OUTPUT_DIR.mkdir(exist_ok=True)
    safe_ma = re.sub(r"[^A-Za-z0-9._-]", "_", product["ma"])
    out_path = OUTPUT_DIR / f"CHIETTINH_{safe_ma}.xlsx"
    shutil.copy(CHIETTINH_FILE, out_path)

    wb = openpyxl.load_workbook(out_path)
    pl1a = wb["PL1A"]
    pl1a["C7"] = product["ma"]
    pl1a["D7"] = product["loai"]
    so_tang = product["so_tang"]
    so_can = product["so_can"]
    if so_tang is not None and so_can is not None:
        pl1a["E7"] = f"CĂN SỐ {so_can} TẦNG {so_tang}"
    gia = product["gia_nn"] if is_foreigner else product["gia_vn"]
    if gia is None:
        raise ValueError(f"Mã {product['ma']} không có giá ({'NN' if is_foreigner else 'VN'})")
    pl1a["C11"] = gia

    if block_value:
        pl1a["C6"] = block_value
    elif product.get("thap"):
        pl1a["C6"] = str(product["thap"])

    # Cập nhật Mã sản phẩm và diện tích trong sheet PL1- Chuẩn (C25 = mã, C26 = diện tích)
    for sn in wb.sheetnames:
        if sn == "PL1A":
            continue
        ws = wb[sn]
        # Các sheet PL1-* thường có ô A25 = "Mã sản phẩm", C25 = giá trị
        try:
            if ws["A25"].value and "Mã sản phẩm" in str(ws["A25"].value):
                ws["C25"] = product["ma"]
                ws["D25"] = product["loai"]
            if ws["A26"].value and "Diện tích" in str(ws["A26"].value):
                if product["dt_tim_tuong"]:
                    ws["C26"] = product["dt_tim_tuong"]
        except Exception:
            pass

    wb.save(out_path)

    # Cố gắng recalc bằng LibreOffice nếu có
    recalc_ok = try_recalc(out_path)
    return out_path, recalc_ok


def try_recalc(path):
    """Mở rồi lưu lại bằng LibreOffice để cập nhật cache công thức."""
    soffice_candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "soffice",
    ]
    for soffice in soffice_candidates:
        try:
            subprocess.run(
                [
                    soffice, "--headless", "--calc",
                    "--convert-to", "xlsx",
                    "--outdir", str(path.parent),
                    str(path),
                ],
                check=True, timeout=60,
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            return True
        except (FileNotFoundError, subprocess.SubprocessError):
            continue
    return False


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tạo Chiết Tính – THE WIN CITY")
        self.geometry("720x520")

        try:
            self.products = load_products()
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
            self.destroy()
            return

        self.codes = sorted(self.products.keys())

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Mã sản phẩm:", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w")
        self.var_code = tk.StringVar()
        self.cb = ttk.Combobox(frm, textvariable=self.var_code, values=self.codes, width=30, font=("Segoe UI", 11))
        self.cb.grid(row=0, column=1, sticky="w", padx=8, pady=4)
        self.cb.bind("<KeyRelease>", self.on_type)
        self.cb.bind("<<ComboboxSelected>>", lambda e: self.refresh_info())

        ttk.Label(frm, text="Đối tượng:").grid(row=1, column=0, sticky="w")
        self.var_kind = tk.StringVar(value="VN")
        kind_frame = ttk.Frame(frm)
        kind_frame.grid(row=1, column=1, sticky="w")
        ttk.Radiobutton(kind_frame, text="Người Việt Nam", variable=self.var_kind, value="VN",
                        command=self.refresh_info).pack(side="left")
        ttk.Radiobutton(kind_frame, text="Người Nước ngoài", variable=self.var_kind, value="NN",
                        command=self.refresh_info).pack(side="left", padx=12)

        ttk.Label(frm, text="Block (tùy chọn):").grid(row=2, column=0, sticky="w")
        self.var_block = tk.StringVar()
        ttk.Entry(frm, textvariable=self.var_block, width=32).grid(row=2, column=1, sticky="w", padx=8, pady=4)

        self.info = tk.Text(frm, height=14, width=80, font=("Consolas", 10), state="disabled")
        self.info.grid(row=3, column=0, columnspan=3, pady=10, sticky="nsew")
        frm.rowconfigure(3, weight=1)
        frm.columnconfigure(1, weight=1)

        btns = ttk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=3, sticky="ew")
        ttk.Button(btns, text="Tạo file Chiết Tính", command=self.do_create).pack(side="left")
        ttk.Button(btns, text="Mở thư mục đầu ra", command=self.open_output).pack(side="left", padx=8)
        ttk.Button(btns, text="Thoát", command=self.destroy).pack(side="right")

        self.status = ttk.Label(frm, text=f"Đã nạp {len(self.products)} mã sản phẩm.", foreground="gray")
        self.status.grid(row=5, column=0, columnspan=3, sticky="w", pady=(8, 0))

    def on_type(self, event):
        s = self.var_code.get().upper()
        filtered = [c for c in self.codes if s in c.upper()]
        self.cb["values"] = filtered or self.codes
        self.refresh_info()

    def refresh_info(self):
        ma = self.var_code.get().strip()
        p = self.products.get(ma)
        self.info.configure(state="normal")
        self.info.delete("1.0", "end")
        if not p:
            self.info.insert("end", "Chưa chọn mã sản phẩm hợp lệ.\n")
        else:
            gia = p["gia_nn"] if self.var_kind.get() == "NN" else p["gia_vn"]
            lines = [
                f"Mã sản phẩm     : {p['ma']}",
                f"Loại            : {p['loai']}",
                f"Tháp            : {p['thap']}",
                f"Tầng / Căn      : {p['so_tang']} / {p['so_can']}",
                f"DT thông thủy   : {p['dt_thong_thuy']} m²",
                f"DT tim tường    : {p['dt_tim_tuong']} m²",
                f"Hướng / View    : {p['huong']} / {p['view']}",
                f"Tình trạng      : {p['tinh_trang']}",
                "",
                f"Giá VN          : {fmt_vnd(p['gia_vn'])}",
                f"Giá Nước ngoài  : {fmt_vnd(p['gia_nn'])}",
                "",
                f">> Giá sẽ điền vào PL1A!C11 = {fmt_vnd(gia)}",
            ]
            self.info.insert("end", "\n".join(lines))
        self.info.configure(state="disabled")

    def do_create(self):
        ma = self.var_code.get().strip()
        p = self.products.get(ma)
        if not p:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng chọn mã sản phẩm hợp lệ.")
            return
        is_nn = self.var_kind.get() == "NN"
        try:
            self.status.config(text="Đang tạo file...", foreground="blue")
            self.update_idletasks()
            out, recalc = create_chiettinh(p, is_nn, self.var_block.get().strip())
        except Exception as e:
            self.status.config(text="Lỗi.", foreground="red")
            messagebox.showerror("Lỗi", str(e))
            return
        note = "" if recalc else "\n\n(Không tìm thấy LibreOffice — khi mở file lần đầu trong Excel, bấm Ctrl+Alt+F9 để tính lại công thức.)"
        self.status.config(text=f"Đã tạo: {out.name}", foreground="green")
        if messagebox.askyesno("Hoàn tất", f"Đã tạo file:\n{out}{note}\n\nMở file ngay?"):
            os.startfile(out)

    def open_output(self):
        OUTPUT_DIR.mkdir(exist_ok=True)
        os.startfile(OUTPUT_DIR)


if __name__ == "__main__":
    App().mainloop()
