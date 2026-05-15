# -*- coding: utf-8 -*-
"""
THE WIN CITY – Công cụ tạo Chiết Tính online (Streamlit).
"""
from __future__ import annotations

import io
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
import streamlit as st

BASE_DIR = Path(__file__).resolve().parent
GIO_HANG_FILE = BASE_DIR / "GIỎ HÀNG LM VAT THE WIN CITY 25.04.2026 (1).xlsx"
CHIETTINH_FILE = BASE_DIR / "CHIETTINH.xlsx"

st.set_page_config(
    page_title="Chiết Tính – THE WIN CITY",
    page_icon="🏙️",
    layout="wide",
)

# ---------- Styling ----------
st.markdown(
    """
    <style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1200px; }
    h1, h2, h3 { font-family: 'Segoe UI', Arial, sans-serif; }
    .hero {
        background: linear-gradient(135deg, #0f2e5c 0%, #1a4d8f 50%, #2563b8 100%);
        padding: 28px 32px; border-radius: 16px; color: white;
        margin-bottom: 24px;
        box-shadow: 0 10px 30px rgba(15,46,92,0.25);
    }
    .hero h1 { color: white; margin: 0; font-size: 2rem; }
    .hero p  { color: #d8e4f5; margin: 6px 0 0 0; font-size: 1rem; }
    .card {
        background: #ffffff; border: 1px solid #e6eaf2;
        border-radius: 14px; padding: 20px 24px;
        box-shadow: 0 4px 14px rgba(15,46,92,0.06);
        margin-bottom: 16px;
    }
    .price-pill {
        display: inline-block; padding: 6px 14px; border-radius: 999px;
        background: #ecf2ff; color: #1a4d8f; font-weight: 600; font-size: 0.95rem;
        margin-right: 8px;
    }
    .price-big {
        font-size: 2rem; font-weight: 700; color: #0f2e5c; margin: 4px 0;
    }
    .muted { color: #6b7280; font-size: 0.9rem; }
    .stDownloadButton button {
        background: linear-gradient(135deg, #1a4d8f, #2563b8) !important;
        color: white !important; border: none !important;
        padding: 12px 24px !important; font-weight: 600 !important;
        border-radius: 10px !important;
    }
    div[data-testid="stMetricValue"] { font-size: 1.4rem; color: #0f2e5c; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------- Helpers ----------
def parse_price(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    digits = re.sub(r"[^\d]", "", str(v))
    return int(digits) if digits else None


@st.cache_data(show_spinner=False)
def load_products():
    wb = openpyxl.load_workbook(GIO_HANG_FILE, data_only=True)
    ws = wb["Sheet1"]
    out = {}
    for r in range(3, ws.max_row + 1):
        ma = ws.cell(r, 2).value
        if not ma:
            continue
        ma = str(ma).strip()
        out[ma] = {
            "ma": ma,
            "loai": ws.cell(r, 3).value,
            "so_can": ws.cell(r, 4).value,
            "so_tang": ws.cell(r, 5).value,
            "thap": ws.cell(r, 6).value,
            "dt_thong_thuy": ws.cell(r, 7).value,
            "dt_tim_tuong": ws.cell(r, 8).value,
            "huong": ws.cell(r, 10).value,
            "view": ws.cell(r, 11).value,
            "gia_vn": parse_price(ws.cell(r, 12).value),
            "gia_nn": parse_price(ws.cell(r, 13).value),
            "tinh_trang": ws.cell(r, 14).value,
        }
    return out


def fmt_vnd(n):
    if n is None:
        return "—"
    return f"{n:,.0f}".replace(",", ".") + " ₫"


DAU_TU_OPTIONS = {
    "Không áp dụng": 0.0,
    "Mua sỉ 2-5 SP (0,5%)": 0.005,
    "Mua sỉ 6-10 SP (0,75%)": 0.0075,
    "Mua sỉ 11-20 SP (1,5%)": 0.015,
    "Mua sỉ ≥21 SP (3%)": 0.03,
}
PTTT_OPTIONS = {
    "PTTT Chuẩn (2%)": 0.02,
    "PTTT Nhanh 50% (5%)": 0.05,
    "PTTT Nhanh 70% (7%)": 0.07,
    "PTTT Nhanh 95% (10%)": 0.10,
    "PTTT Vay (0% – có voucher 20tr)": 0.0,
    "Không áp dụng": 0.0,
}


def compute_breakdown(gia: int, pol: dict):
    """Tái hiện công thức trong CHIETTINH.xlsx để preview."""
    e12 = pol["e12"]; e13 = pol["e13"]; e14 = pol["e14"]; c15 = pol["c15"]
    c12 = round(gia * e12)
    c13 = round((gia - c12) * e13)
    c14 = round((gia - c12 - c13) * e14)
    c15v = int(c15)
    gia_truoc_vat = gia - c12 - c13 - c14 - c15v
    vat = round(gia_truoc_vat * 0.10)
    tong_co_vat = gia_truoc_vat + vat
    phi_bao_tri = round(gia_truoc_vat * 0.02)
    tong_cuoi = tong_co_vat + phi_bao_tri
    return {
        "Giá CĐT (trước VAT)": gia,
        'CK "Phi Mã Đón Đầu"': -c12,
        'CK "Đầu Tư Gia Tăng"': -c13,
        "CK theo PTTT": -c14,
        "CK khác (voucher)": -c15v,
        "Giá HĐMB trước VAT": gia_truoc_vat,
        "VAT (10%)": vat,
        "Tổng HĐMB (gồm VAT)": tong_co_vat,
        "Phí bảo trì (2%)": phi_bao_tri,
        "TỔNG GIÁ TRỊ HỢP ĐỒNG": tong_cuoi,
    }


def try_recalc(path: Path) -> tuple[bool, str]:
    """Tính lại công thức bằng LibreOffice. Trả về (ok, log)."""
    candidates = [
        "/usr/bin/libreoffice", "/usr/bin/soffice",
        "libreoffice", "soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
    ]
    out_dir = path.parent / "recalc_out"
    out_dir.mkdir(exist_ok=True)
    last_err = ""
    for soffice in candidates:
        try:
            r = subprocess.run(
                [soffice, "--headless", "--calc",
                 "--convert-to", "xlsx",
                 "--outdir", str(out_dir), str(path)],
                check=True, timeout=120,
                capture_output=True, text=True,
            )
            converted = out_dir / path.name
            if converted.exists():
                shutil.copy(converted, path)
                return True, f"ok via {soffice}"
            last_err = f"{soffice}: converted file not found. stdout={r.stdout[:200]} stderr={r.stderr[:200]}"
        except FileNotFoundError:
            last_err = f"{soffice}: not found"
            continue
        except subprocess.CalledProcessError as e:
            last_err = f"{soffice}: rc={e.returncode} stderr={e.stderr[:300] if e.stderr else ''}"
            continue
        except subprocess.SubprocessError as e:
            last_err = f"{soffice}: {e}"
            continue
    return False, last_err


def fallback_compute_pl1a(wb, gia: int, dt_tim_tuong, pol: dict):
    """Khi không có LibreOffice: tính trực tiếp và ghi đè value cho các ô chính
    để file mở ra đã thấy số ngay (không cần Excel recalc)."""
    pl1a = wb["PL1A"]
    e12 = pol["e12"]; e13 = pol["e13"]; e14 = pol["e14"]; c15 = int(pol["c15"])
    c12 = round(gia * e12)
    c13 = round((gia - c12) * e13)
    c14 = round((gia - c12 - c13) * e14)
    c22 = gia - c12 - c13 - c14 - c15
    pl1a["C12"] = c12
    pl1a["C13"] = c13
    pl1a["C14"] = c14
    if c15:
        pl1a["C15"] = c15
    pl1a["C22"] = c22

    # PL1-Chuẩn (và các sheet phương thức): C30=C22, C31, C32, C33, C34, C35
    dt = float(dt_tim_tuong) if dt_tim_tuong else 0
    for sn in wb.sheetnames:
        if sn == "PL1A":
            continue
        ws = wb[sn]
        # Nhận diện các ô cần điền dựa trên label cột B
        c26 = ws["C26"].value
        if isinstance(c26, (int, float)):
            dt_local = float(c26)
        else:
            dt_local = dt
        # C30 = Giá HĐMB trước VAT
        ws["C30"] = c22
        # C31 = giá đất được trừ
        c31 = round(dt_local * 1000079)
        ws["C31"] = c31
        # C32 = VAT
        c32 = round((c22 - c31) * 0.10)
        ws["C32"] = c32
        # C33 = C30 + C32
        ws["C33"] = c22 + c32
        # C34 = phí bảo trì = 2% * C30
        c34 = round(c22 * 0.02)
        ws["C34"] = c34
        # C35 = C33 + C34
        ws["C35"] = c22 + c32 + c34


def build_chiettinh(product: dict, is_foreigner: bool, block: str, pol: dict) -> tuple[bytes, bool, str]:
    gia = product["gia_nn"] if is_foreigner else product["gia_vn"]
    if gia is None:
        raise ValueError("Mã sản phẩm này chưa có giá.")

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", product["ma"])
        out_path = td / f"CHIETTINH_{safe}.xlsx"
        shutil.copy(CHIETTINH_FILE, out_path)

        wb = openpyxl.load_workbook(out_path)
        pl1a = wb["PL1A"]
        pl1a["C7"] = product["ma"]
        pl1a["D7"] = product["loai"]
        if product["so_tang"] is not None and product["so_can"] is not None:
            pl1a["E7"] = f"CĂN SỐ {product['so_can']} TẦNG {product['so_tang']}"
        pl1a["C11"] = gia
        # Ghi tỉ lệ chiết khấu vào các ô input của template
        pl1a["E12"] = pol["e12"]
        pl1a["E13"] = pol["e13"]
        pl1a["E14"] = pol["e14"]
        if pol["c15"]:
            pl1a["C15"] = int(pol["c15"])
        if block:
            pl1a["C6"] = block
        elif product.get("thap"):
            pl1a["C6"] = str(product["thap"])

        for sn in wb.sheetnames:
            if sn == "PL1A":
                continue
            ws = wb[sn]
            try:
                if ws["A25"].value and "Mã sản phẩm" in str(ws["A25"].value):
                    ws["C25"] = product["ma"]
                    ws["D25"] = product["loai"]
                if ws["A26"].value and "Diện tích" in str(ws["A26"].value):
                    if product["dt_tim_tuong"]:
                        ws["C26"] = product["dt_tim_tuong"]
            except Exception:
                pass

        # Bắt Excel tính lại toàn bộ công thức khi mở file
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(calcId=191241, fullCalcOnLoad=True, calcMode="auto")
        for sn in wb.sheetnames:
            ws = wb[sn]
            ws.force_formula_recalculation = True if hasattr(ws, "force_formula_recalculation") else None

        wb.save(out_path)
        ok, log = try_recalc(out_path)
        if not ok:
            # Không có LibreOffice → ghi đè value cho các ô quan trọng
            wb2 = openpyxl.load_workbook(out_path)
            fallback_compute_pl1a(wb2, gia, product.get("dt_tim_tuong"), pol)
            wb2.save(out_path)
        data = out_path.read_bytes()
        return data, ok, log


# ---------- UI ----------
st.markdown(
    """
    <div class="hero">
      <h1>🏙️ Công cụ Chiết Tính – THE WIN CITY</h1>
      <p>Chọn mã sản phẩm → tự động tra giá từ Giỏ hàng → xuất file Chiết Tính kèm tiến độ thanh toán.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

try:
    products = load_products()
except FileNotFoundError as e:
    st.error(f"❌ Thiếu file dữ liệu: {e}")
    st.stop()

codes = sorted(products.keys())

col_l, col_r = st.columns([1, 1.2], gap="large")

with col_l:
    st.markdown("### 1. Chọn căn")
    code = st.selectbox(
        "Mã sản phẩm",
        options=codes,
        index=0,
        help=f"Có {len(codes)} mã căn trong file Giỏ hàng.",
    )
    kind = st.radio(
        "Đối tượng khách hàng",
        options=["Người Việt Nam", "Người Nước ngoài"],
        horizontal=True,
    )
    block = st.text_input("Block (tùy chọn)", placeholder="VD: C1.3 – để trống dùng giá trị mặc định")
    st.caption("Giá và mọi công thức trong file CHIETTINH sẽ tự cập nhật theo lựa chọn.")

product = products[code]
is_nn = kind == "Người Nước ngoài"
gia = product["gia_nn"] if is_nn else product["gia_vn"]

with col_r:
    st.markdown("### 2. Thông tin căn")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown(
        f'<span class="price-pill">{product["loai"] or "—"}</span>'
        f'<span class="price-pill">Tháp {product["thap"]}</span>'
        f'<span class="price-pill">Tầng {product["so_tang"]} – Căn {product["so_can"]}</span>',
        unsafe_allow_html=True,
    )
    st.markdown(f'<div class="price-big">{fmt_vnd(gia)}</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="muted">Giá CĐT trước VAT · {kind}</div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    c1.metric("DT thông thủy", f"{product['dt_thong_thuy']} m²" if product['dt_thong_thuy'] else "—")
    c2.metric("DT tim tường", f"{product['dt_tim_tuong']} m²" if product['dt_tim_tuong'] else "—")
    c3.metric("Hướng", product["huong"] or "—")
    st.markdown(f'<div class="muted">View: {product["view"] or "—"} · Tình trạng: {product["tinh_trang"] or "—"}</div>',
                unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown("### 3. Chính sách / Chiết khấu áp dụng")
st.caption("Bỏ chọn để ngưng áp dụng. Có thể điều chỉnh tỉ lệ tùy ý.")

pc1, pc2 = st.columns(2)
with pc1:
    apply_phi_ma = st.checkbox('Chương trình "Phi Mã Đón Đầu"', value=True)
    phi_ma_pct = st.number_input("Tỷ lệ CK Phi Mã Đón Đầu (%)", min_value=0.0, max_value=20.0,
                                  value=2.0, step=0.1, disabled=not apply_phi_ma)

    apply_dau_tu = st.checkbox('Chương trình "Đầu Tư Gia Tăng" (mua sỉ)', value=False)
    dau_tu_label = st.selectbox("Mức mua sỉ", list(DAU_TU_OPTIONS.keys()),
                                 index=0, disabled=not apply_dau_tu)

with pc2:
    apply_pttt = st.checkbox("Chiết khấu theo Phương thức thanh toán", value=True)
    pttt_label = st.selectbox("Phương thức thanh toán",
                               list(PTTT_OPTIONS.keys()), index=0, disabled=not apply_pttt)

    is_vay = apply_pttt and pttt_label.startswith("PTTT Vay")
    apply_voucher = st.checkbox("Voucher 20 triệu (chỉ áp dụng PTTT Vay)",
                                 value=is_vay, disabled=not is_vay)
    voucher_amt = st.number_input("Số tiền voucher (VNĐ)", min_value=0,
                                   value=20_000_000 if apply_voucher else 0,
                                   step=1_000_000, disabled=not apply_voucher)

policies = {
    "e12": (phi_ma_pct / 100.0) if apply_phi_ma else 0.0,
    "e13": DAU_TU_OPTIONS[dau_tu_label] if apply_dau_tu else 0.0,
    "e14": PTTT_OPTIONS[pttt_label] if apply_pttt else 0.0,
    "c15": voucher_amt if (is_vay and apply_voucher) else 0,
}

with st.expander("Các chương trình quà tặng (không trừ vào giá)"):
    st.checkbox('Cam kết dòng tiền thuê 3%/năm (24 tháng) – chỉ C1.3', value=False, key="cs_thue")
    st.checkbox('Chương trình "Chìa Khóa Trao Tay – Xách Vali ở Ngay"', value=False, key="cs_chiakhoa")
    st.checkbox('Chương trình "Nhà mới trao tay"', value=False, key="cs_nhamoi")
    st.caption("Các chương trình này được ghi nhận trong hợp đồng nhưng KHÔNG trừ vào giá HĐMB.")

st.markdown("### 4. Bảng tính nhanh")
if gia:
    bd = compute_breakdown(gia, policies)
    cols = st.columns(5)
    items = list(bd.items())
    for i, (k, v) in enumerate(items):
        with cols[i % 5]:
            label = k
            display = fmt_vnd(abs(v)) if v >= 0 else f"−{fmt_vnd(abs(v))}"
            st.metric(label, display)
    st.caption("Cập nhật theo chính sách bạn chọn ở mục 3.")
else:
    st.warning("Mã này chưa có giá trong file Giỏ hàng.")

st.markdown("### 5. Xuất file Chiết Tính")
go = st.button("⚙️ Tạo file Chiết Tính", type="primary", use_container_width=True, disabled=gia is None)

if go and gia:
    with st.spinner("Đang điền dữ liệu vào template và tính công thức..."):
        try:
            data, recalc_ok, recalc_log = build_chiettinh(product, is_nn, block.strip(), policies)
        except Exception as e:
            st.error(f"Lỗi: {e}")
            st.stop()

    safe = re.sub(r"[^A-Za-z0-9._-]", "_", product["ma"])
    suffix = "NN" if is_nn else "VN"
    st.success("✅ Đã tạo xong file. Bấm tải xuống bên dưới.")
    if recalc_ok:
        st.caption(f"✓ Đã tính lại công thức bằng LibreOffice ({recalc_log})")
    else:
        st.info("ℹ️ Không có LibreOffice trên server — các ô chính đã được tính sẵn bằng Python "
                "(Giá HĐMB, VAT, Phí bảo trì, Tổng giá). "
                "Khi mở file trong Excel, bấm **Ctrl+Alt+F9** để cập nhật toàn bộ công thức còn lại.")
        with st.expander("Chi tiết lỗi recalc"):
            st.code(recalc_log)
    st.download_button(
        label=f"⬇️ Tải CHIETTINH_{safe}_{suffix}.xlsx",
        data=data,
        file_name=f"CHIETTINH_{safe}_{suffix}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.markdown("---")
st.caption("© THE WIN CITY · Dữ liệu lấy từ file Giỏ hàng nội bộ · Bản công cụ chiết tính tự động.")
